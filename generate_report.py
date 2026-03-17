import mysql.connector
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import date, datetime

# Connect to database
conn = mysql.connector.connect(
    host="localhost",
    user="root",
    password="Anu@2005",
    database="attendance_db"
)
cursor = conn.cursor()

# Get all attendance data
cursor.execute("""
    SELECT e.name, e.department, e.email,
           a.date, a.punch_in, a.punch_out, a.status
    FROM attendance a
    JOIN employees e ON a.employee_id = e.id
    ORDER BY a.date DESC
""")
rows = cursor.fetchall()

# Get absent employees today
today = date.today()
cursor.execute("""
    SELECT e.name, e.department, e.email
    FROM employees e
    WHERE e.id NOT IN (
        SELECT employee_id FROM attendance WHERE date=%s
    )
""", (today,))
absent_rows = cursor.fetchall()
conn.close()

# Create Excel workbook
wb = Workbook()

# ── Sheet 1: Attendance Report ──
ws1 = wb.active
ws1.title = "Attendance Report"

# Title
ws1.merge_cells('A1:G1')
ws1['A1'] = f"Attendance Report - Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}"
ws1['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws1['A1'].fill = PatternFill("solid", fgColor="4A90E2")
ws1['A1'].alignment = Alignment(horizontal="center")

# Headers
headers = ["Name", "Department", "Email", "Date", "Punch In", "Punch Out", "Status"]
for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=2, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2C3E50")
    cell.alignment = Alignment(horizontal="center")

# Data rows
for row_idx, row in enumerate(rows, 3):
    for col_idx, value in enumerate(row, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=str(value) if value else "-")
        cell.alignment = Alignment(horizontal="center")
        # Alternate row colors
        if row_idx % 2 == 0:
            cell.fill = PatternFill("solid", fgColor="EBF5FB")

# Status color coding
for row_idx in range(3, len(rows) + 3):
    status_cell = ws1.cell(row=row_idx, column=7)
    if status_cell.value == "Present":
        status_cell.font = Font(color="27AE60", bold=True)
    else:
        status_cell.font = Font(color="E74C3C", bold=True)

# Column widths
column_widths = [20, 20, 30, 15, 12, 12, 12]
for col, width in enumerate(column_widths, 1):
    ws1.column_dimensions[get_column_letter(col)].width = width

# ── Sheet 2: Absent Employees ──
ws2 = wb.create_sheet("Absent Today")

ws2.merge_cells('A1:C1')
ws2['A1'] = f"Absent Employees - {today}"
ws2['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws2['A1'].fill = PatternFill("solid", fgColor="E74C3C")
ws2['A1'].alignment = Alignment(horizontal="center")

absent_headers = ["Name", "Department", "Email"]
for col, header in enumerate(absent_headers, 1):
    cell = ws2.cell(row=2, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2C3E50")
    cell.alignment = Alignment(horizontal="center")

if absent_rows:
    for row_idx, row in enumerate(absent_rows, 3):
        for col_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=str(value))
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(color="E74C3C")
else:
    ws2.cell(row=3, column=1, value="All employees present today! 🎉")
    ws2['A3'].font = Font(color="27AE60", bold=True)

for col, width in enumerate([20, 20, 30], 1):
    ws2.column_dimensions[get_column_letter(col)].width = width

# ── Sheet 3: Summary ──
ws3 = wb.create_sheet("Summary")

ws3.merge_cells('A1:B1')
ws3['A1'] = "Attendance Summary"
ws3['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws3['A1'].fill = PatternFill("solid", fgColor="27AE60")
ws3['A1'].alignment = Alignment(horizontal="center")

summary_data = [
    ("Report Date", str(today)),
    ("Total Employees", len(rows)),
    ("Present Today", len(rows) - len(absent_rows)),
    ("Absent Today", len(absent_rows)),
    ("Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
]

for row_idx, (key, value) in enumerate(summary_data, 2):
    ws3.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
    ws3.cell(row=row_idx, column=2, value=str(value))

ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 25

# Save file
filename = f"attendance_report_{today}.xlsx"
wb.save(filename)
print(f"✅ Excel report saved as: {filename}")
print(f"📊 Total records: {len(rows)}")
print(f"❌ Absent today: {len(absent_rows)}")
